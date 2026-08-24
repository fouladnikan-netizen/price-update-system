#!/usr/bin/env python3
"""Parse Final Product.xlsx Sheet1 into excel-catalog-rows.json (exact file rows)."""
import json
import os
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), "..")
EXCEL = os.path.join(ROOT, "data", "products", "Final Product.xlsx")
OUT = os.path.join(ROOT, "apps", "web", "src", "mock", "excel-catalog-rows.json")


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text in ("None", "NULL", "null"):
        return ""
    return text


wb = load_workbook(EXCEL, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = [clean(h) for h in rows[0]]


def col(*names):
    for i, h in enumerate(header):
        for n in names:
            if n and n in h:
                return i
    return None


c_cat = col("دسته")
c_name = col("نام کالا", "نام")
c_type = col("نوع")
c_dim = col("ابعاد", "طول")
c_size = col("سایز")
c_weight = col("وزن")
c_unit = col("واحد")
c_sched = col("رده")
c_thick = None
for i, h in enumerate(header):
    if h == "ضخامت":
        c_thick = i
        break

items = []
for row in rows[1:]:
    if not row:
        continue
    category = clean(row[c_cat]) if c_cat is not None else ""
    name = clean(row[c_name]) if c_name is not None else ""
    if not category or not name:
        continue
    items.append(
        {
            "category": category,
            "name": name,
            "kind": clean(row[c_type]) if c_type is not None else "",
            "dimensions": clean(row[c_dim]) if c_dim is not None else "",
            "size": clean(row[c_size]) if c_size is not None else "",
            "weight": clean(row[c_weight]) if c_weight is not None else "",
            "unit": clean(row[c_unit]) if c_unit is not None else "کیلوگرم",
            "pipeClass": clean(row[c_sched]) if c_sched is not None else "",
            "thickness": clean(row[c_thick]) if c_thick is not None else "",
        }
    )

wb.close()
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(items, f, ensure_ascii=False, indent=2)
    f.write("\n")

from collections import Counter

print(json.dumps({"out": OUT, "count": len(items), "byCategory": dict(Counter(i["category"] for i in items))}, ensure_ascii=False, indent=2))
